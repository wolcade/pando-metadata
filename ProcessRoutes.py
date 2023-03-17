import csv
from decimal import *
import piexif
from PIL import Image, ExifTags
import os
import argparse
import time
from pathlib import Path
import shutil
from openpyxl import load_workbook
import csv


PRECISION_TEN = Decimal('0.0000000001')


def getRouteFolders(folder_path):
    files = os.listdir(folder_path)
    return [name for name in files if os.path.isdir(os.path.join(folder_path, name))]


def getFiles(folder_path):
    files = os.listdir(folder_path)
    return [name for name in files if os.path.isfile(os.path.join(folder_path, name))]


def createGPSCSVFromExcel(excelFile, routeFolderPath):
    wb = load_workbook(filename=excelFile)
    ws = wb.worksheets[0]

    # location data in column B starts at row 18
    locations = [cell.value for cell in ws['B']][17:]

    with open(os.path.join(routeFolderPath, 'GPSData.csv'), 'w', newline='') as csvFile:
        writer = csv.writer(csvFile)
        writer.writerow(['Location', 'Rec Lat', 'Rec Lon'])
        # Grab the rec lat and rec lon columns starting at row 18
        locationDataRows = [row[-2:] for row in ws.iter_rows(min_row=18)]

        for count, location in enumerate(locations):
            writer.writerow([location, locationDataRows[count]
                            [0].value, locationDataRows[count][1].value])


def readCSVToDict(csvPath):
    # {[locationKey]: {lat: number, lon: number}}
    gpsLocationDict = {}
    with open(csvPath, mode='r') as file:
        csvFile = csv.DictReader(file)
        for row in csvFile:
            gpsLocationDict[row["Location"]] = {
                "lat": row['Rec Lat'], "lon": row["Rec Lon"]}
    return gpsLocationDict


def processRoutes(folder_path):
    routes = getRouteFolders(folder_path)

    for route in routes:
        routeFolderPath = os.path.join(folder_path, route)
        directories = getRouteFolders(routeFolderPath)
        excelFiles = list(filter(lambda file: ".xls" in file,
                          getFiles(routeFolderPath)))

        if len(excelFiles) > 0:
            createGPSCSVFromExcel(os.path.join(
                routeFolderPath, excelFiles[0]), routeFolderPath)

        # create new folder for images if does not already exist
        newFolder = os.path.join(folder_path, route) + "/imagesWithGPS"
        Path(newFolder).mkdir(parents=True, exist_ok=True)

        images = getFiles(os.path.join(folder_path, route, "Images"))

        # copy images to new folder
        for image in images:
            shutil.copy2(os.path.join(folder_path, route,
                         "Images", image), newFolder)

        gpsLocationDictionary = readCSVToDict(
            os.path.join(routeFolderPath, 'GPSData.csv'))

        directory = os.fsencode(newFolder)
        files = os.listdir(directory)
        fileList = [os.fsdecode(f) for f in files]
        print('Total files in directory... ', len(fileList))

        imageCount = 0

        success = []
        failure = []
        missing = []

        for file in fileList:
            img_path = f'{newFolder}/{file}'

            if file.endswith('.jpg') and not (file.startswith('EndSlate') or file.startswith('OpenSlate')):
                location = file.split('-')[0]

                #print(f'Processing Location {location} from csv')

                # Grab lat/lon from csv data
                gpsData = gpsLocationDictionary[location.upper()]

                setGPSData(img_path, img_path, gpsData, location)

                imageCount += 1

                validationStatus = validateGPSData(
                    img_path, gpsData, location)
                if validationStatus == 1:
                    success.append(location)
                if validationStatus == 0:
                    missing.append(location)
                if validationStatus == -1:
                    failure.append(location)
            else:
                print('skipping ', file)

        print(
            f'Completed Processing {folder_path}. {imageCount} images processed.\n')
        print(
            f'\nVALIDATION SUMMARY\n\nSuccess:\n{success}\n{len(success)} files\n\nFailure\n{failure}\n{len(failure)} files\n\nMissing GPS Data\n{missing}\n{len(missing)} files')


def validateGPSData(new_img_path, gpsData, location):

    if gpsData["lat"] != '' and gpsData["lon"] != '':
        im = Image.open(new_img_path)
        mTime = os.path.getmtime(new_img_path)
        modifiedTime = time.ctime(mTime)

        exif_dict = piexif.load(im.info["exif"])

        newGPSLat = exif_dict["GPS"][piexif.GPSIFD.GPSLatitude]
        decimalDegreesLat = convertDMSToDD(newGPSLat)

        newGPSLon = exif_dict["GPS"][piexif.GPSIFD.GPSLongitude]
        decimalDegreesLon = convertDMSToDD(newGPSLon) * -1
        #print(f"FROM IMAGE {newGPSLat}, {newGPSLon}")

        #print("FROM CSV ", gpsData["lat"], gpsData["lon"])
        gpsDataLat = Decimal(gpsData["lat"])
        gpsDataLon = Decimal(gpsData["lon"])

        
        #print(f"    Reading Image {location} modified at {modifiedTime}")
        #print(f"    DMS Coordinates in Image: {newGPSLat}, {newGPSLon}")

        """ print(
            f"    converted exif GPS lat {decimalDegreesLat} equals csv lat {gpsDataLat}: {decimalDegreesLat == gpsDataLat}")
        print(
            f"    converted exif GPS lon {decimalDegreesLon} equals csv lon {gpsDataLon}: {decimalDegreesLon == gpsDataLon}")
        print(
            f"    DMS Coordinates in Bridge Format: {convertDegreesToDMSBridge(decimalDegreesLat)}, {convertDegreesToDMSBridge(decimalDegreesLon)}")

        print('') """

        if decimalDegreesLat == gpsDataLat and decimalDegreesLon == gpsDataLon:
            return 1
        else:
            return -1

    else:
        print(f"    missing gps data for {location}")
        return 0


def convertDegreesToDMSBridge(decimalDegrees):
    # Set precision to at least 10 decimal places. Need to validate further
    c = getcontext()
    c.prec = 20

    decimal = Decimal(decimalDegrees)
    degrees = abs(int(decimal))
    minutes = (decimal % 1) * 60
    seconds = (minutes % 1) * 60

    minutes = abs(minutes)
    return (degrees, minutes)


def convertDMSToDD(dmsTuple):

    degrees = Decimal(dmsTuple[0][0])
    minutes = Decimal(dmsTuple[1][0])
    seconds = Decimal(dmsTuple[2][0]) / Decimal(dmsTuple[2][1])

    decimalDegrees = Decimal(
        degrees) + Decimal(minutes / 60) + Decimal(seconds / (60 * 60))

    return Decimal(decimalDegrees)


def convertDegreesToDMS(decimalDegrees):
    # Set precision to at least 10 decimal places. Need to validate further
    c = getcontext()
    c.prec = 20

    decimal = Decimal(decimalDegrees)
    degrees = abs(int(decimal))
    minutes = (decimal % 1) * 60
    seconds = (minutes % 1) * 60

    minutes = abs(int(minutes))
    (seconds_num, seconds_denominator) = abs((seconds)).as_integer_ratio()
    return ((degrees, 1), (minutes, 1), (seconds_num, seconds_denominator))


def setGPSData(img_path, new_img_path, gpsData, location):
    # For each image, load the exif
    # set the tags for GPS
    # insert new tags into image
    # save image

    im = Image.open(img_path)
    exif_dict = piexif.load(img_path)

    if gpsData["lat"] != '' and gpsData["lon"] != '':
        exif_dict["GPS"][piexif.GPSIFD.GPSLatitude] = convertDegreesToDMS(
            gpsData["lat"])
        exif_dict["GPS"][piexif.GPSIFD.GPSLongitude] = convertDegreesToDMS(
            gpsData["lon"])
        exif_dict["GPS"][piexif.GPSIFD.GPSLongitudeRef] = "W"
        exif_dict["GPS"][piexif.GPSIFD.GPSLatitudeRef] = "N"

        """ print(
            f'    Set GPSLatitude: {gpsData["lat"]} to {exif_dict["GPS"][piexif.GPSIFD.GPSLatitude]}')
        print(
            f'    Set GPSLongitude: {gpsData["lon"]} to {exif_dict["GPS"][piexif.GPSIFD.GPSLongitude]}')
        print(
            f'    Set GPSLatitudeRef to {exif_dict["GPS"][piexif.GPSIFD.GPSLatitudeRef]}')
        print(
            f'    Set GPSLongitudeRef to {exif_dict["GPS"][piexif.GPSIFD.GPSLongitudeRef]}')
        print('') """

    else:
        print(f'    Missing Coordinates for location {location}\n')

    exif_bytes = piexif.dump(exif_dict)
    piexif.insert(exif_bytes, img_path)
    #print(new_img_path + ' saved successfully')
    #print('')


def main():
    c = getcontext()
    c.prec = 20

    # Args: image directory, csv file, writeFlag, verbose, validate
    parser = argparse.ArgumentParser(
        description='Insert GPS coordinates into images')
    parser.add_argument('--imageDirectory', type=str, required=True)

    args = parser.parse_args()

    folder_path = args.imageDirectory
    if not os.path.exists(folder_path):
        raise ValueError(f'Invalid Path. Cannot find {folder_path}.')

    processRoutes(folder_path)


if __name__ == "__main__":
    main()
